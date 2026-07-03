#!/usr/bin/env python3
"""Fill SGMI double-entry traceability matrix from user stories."""
from __future__ import annotations

import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = Path(r"c:\Users\INTEL\Downloads\PLANTILLA - MATRIZ DE DOBLE ENTRADA.xlsx")
OUTPUT_PROJECT = ROOT / "docs" / "03_implementacion" / "MATRIZ-DOBLE-ENTRADA-SGMI.xlsx"
OUTPUT_DOWNLOADS = Path(
    r"c:\Users\INTEL\Downloads\MATRIZ-DOBLE-ENTRADA-SGMI.xlsx"
)
COMMITS_JSON = ROOT / "docs" / "03_implementacion" / "trazabilidad" / "commits.json"

COMMITS = {
    "init": "f56b168",
    "frontend": "6877d83",
    "prompt_fe_v1": "f05c90a",
    "prompt_fe_v2": "ee948b0",
    "backend": "5a891bd",
    "merge": "7584a84",
    "integraciones": "71144a5",
}

PROMPTS = {
    "R-001": {
        "id": "P-R-001",
        "file": "prompts/01_requisitos/R-001_elicitacion_ambiguedades_v1.1.md",
        "version": "v1.1",
        "technique": "CoT",
        "expected": "Elicitar requisitos y resolver ambigüedades del MVP",
    },
    "R-002": {
        "id": "P-R-002",
        "file": "prompts/01_requisitos/R-002_historias_usuario_v1.md",
        "version": "v1",
        "technique": "Role + Gherkin",
        "expected": "Generar 26 historias de usuario con criterios Gherkin",
    },
    "D-001": {
        "id": "P-D-001",
        "file": "prompts/02_diseno/D-001_modelo_datos_v1.md",
        "version": "v1",
        "technique": "ER + SQL",
        "expected": "Modelo ER normalizado y migración MySQL SGMI",
    },
    "D-002": {
        "id": "P-D-002",
        "file": "prompts/02_diseno/D-002_decision_arquitectura_v1.md",
        "version": "v1",
        "technique": "ADR",
        "expected": "Decisiones arquitectónicas Laravel + Vue + MySQL",
    },
    "D-003-v1": {
        "id": "P-D-003-v1",
        "file": "prompts/02_diseno/D-003 - Promtp de Diseño FrontEnd (Stich) - v1",
        "version": "v1",
        "technique": "Few-shot + diseño UI",
        "expected": "Mockups HTML/CSS pantallas prioritarias MOD-DOC y MOD-DASH",
    },
    "D-003-v2": {
        "id": "P-D-003-v2",
        "file": "prompts/02_diseno/D-003 - Promtp de Diseño FrontEnd (Stich) - v2",
        "version": "v2",
        "technique": "Few-shot + diseño UI",
        "expected": "Landing municipal + shell SGMI + 6 pantallas MVP",
    },
}

HISTORIAS = [
    {
        "hu": "HU-SEC-01",
        "title": "Autenticación local segura",
        "module": "NÚCLEO",
        "rf": "RF-NC-01, RF-NC-02",
        "status": "Implementado",
        "location": "app/Http/Controllers/Api/AuthController.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/NucleoTest.php",
        "commit": COMMITS["backend"],
        "prompt": "D-001",
        "prompt_text": "Implementar autenticación local con bloqueo 5 intentos y política de contraseña",
        "notes": "Login sesión; PasswordSegura; bcrypt",
    },
    {
        "hu": "HU-SEC-02",
        "title": "Permisos por rol y organigrama",
        "module": "NÚCLEO",
        "rf": "RF-NC-03, RF-NC-04, RF-NC-10",
        "status": "Implementado",
        "location": "app/Services/Core/AccesoService.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/NucleoTest.php",
        "commit": COMMITS["backend"],
        "prompt": "D-001",
        "prompt_text": "RBAC por organigrama; rol VISTA_EJECUTIVA; menú por permisos",
        "notes": "EnsurePermission middleware; MenuService",
    },
    {
        "hu": "HU-SEC-03",
        "title": "Auditoría de operaciones",
        "module": "NÚCLEO",
        "rf": "RF-NC-05",
        "status": "Implementado",
        "location": "app/Services/Core/AuditoriaService.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/NucleoTest.php",
        "commit": COMMITS["backend"],
        "prompt": "D-001",
        "prompt_text": "Servicio append-only de auditoría transversal",
        "notes": "Logs en operaciones CRUD de dominio",
    },
    {
        "hu": "HU-SEC-04",
        "title": "Consulta auditoría OCI",
        "module": "NÚCLEO",
        "rf": "RF-NC-06",
        "status": "Implementado",
        "location": "app/Http/Controllers/Api/AuditoriaController.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/NucleoTest.php",
        "commit": COMMITS["backend"],
        "prompt": "D-001",
        "prompt_text": "Consulta paginada solo lectura para rol AUDITOR_OCI",
        "notes": "AuditoriaPage.vue; export CSV",
    },
    {
        "hu": "HU-SEC-05",
        "title": "Administración de usuarios y traslados",
        "module": "NÚCLEO",
        "rf": "RF-NC-07, RF-NC-09",
        "status": "Implementado",
        "location": "app/Services/Core/UsuarioAdminService.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/NucleoTest.php",
        "commit": COMMITS["backend"],
        "prompt": "D-001",
        "prompt_text": "CRUD usuarios UTIS con historial de traslados",
        "notes": "UsuarioTraslado; UsuariosPage.vue",
    },
    {
        "hu": "HU-ORG-01",
        "title": "Catálogo de unidades (gerencia real)",
        "module": "NÚCLEO",
        "rf": "RF-NC-08",
        "status": "Parcial",
        "location": "app/Http/Controllers/Api/UnidadController.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/NucleoTest.php",
        "commit": COMMITS["backend"],
        "prompt": "R-002",
        "prompt_text": "Árbol jerárquico; excluir comités de derivación",
        "notes": "Seeder demo ~23 unidades; organigrama completo 61 pendiente",
    },
    {
        "hu": "HU-DOC-01",
        "title": "Registro expediente con numeración por tipo",
        "module": "MOD-DOC",
        "rf": "RF-DOC-01, RF-DOC-02",
        "status": "Implementado",
        "location": "app/Services/Documentaria/ExpedienteService.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/ExpedienteDocumentalTest.php",
        "commit": COMMITS["backend"],
        "prompt": "D-001",
        "prompt_text": "Numeración {prefijo}-{año}-{secuencia} por tipo documental",
        "notes": "RegistroExpedientePage.vue",
    },
    {
        "hu": "HU-DOC-02",
        "title": "Derivación y devolución",
        "module": "MOD-DOC",
        "rf": "RF-DOC-03, RF-DOC-04, RF-DOC-05",
        "status": "Implementado",
        "location": "app/Services/Documentaria/ExpedienteService.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/ExpedienteDocumentalTest.php",
        "commit": COMMITS["backend"],
        "prompt": "R-002",
        "prompt_text": "Derivación libre; devolución automática al remitente",
        "notes": "BandejaPendientesPage.vue; ConstanciaService",
    },
    {
        "hu": "HU-DOC-03",
        "title": "Trazabilidad e historial",
        "module": "MOD-DOC",
        "rf": "RF-DOC-06",
        "status": "Implementado",
        "location": "app/Services/Documentaria/ExpedienteHistorialService.php",
        "version": "v1.0",
        "modified": "2026-06-26",
        "test": "tests/Feature/ExpedienteDocumentalTest.php",
        "commit": COMMITS["integraciones"],
        "prompt": "D-003-v2",
        "prompt_text": "Línea de tiempo vertical con permanencia por oficina",
        "notes": "TrazabilidadExpedientePage.vue; TimelineNode.vue",
    },
    {
        "hu": "HU-DOC-04",
        "title": "Bandeja de pendientes",
        "module": "MOD-DOC",
        "rf": "RF-DOC-07, RF-DOC-08",
        "status": "Implementado",
        "location": "app/Http/Controllers/Api/ExpedienteController.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/ExpedienteDocumentalTest.php",
        "commit": COMMITS["backend"],
        "prompt": "D-003-v2",
        "prompt_text": "Bandeja por unidad con filtros prioridad/tipo/antigüedad",
        "notes": "BandejaPendientesPage.vue",
    },
    {
        "hu": "HU-DOC-05",
        "title": "Adjuntos digitalizados",
        "module": "MOD-DOC",
        "rf": "RF-DOC-09",
        "status": "Implementado",
        "location": "app/Http/Controllers/Api/ExpedienteAdjuntoController.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/ExpedienteDocumentalTest.php",
        "commit": COMMITS["backend"],
        "prompt": "D-001",
        "prompt_text": "Adjuntos en registro; rechazo de ejecutables",
        "notes": "FileDropzone.vue; AnexoList.vue",
    },
    {
        "hu": "HU-DOC-06",
        "title": "Búsqueda de expedientes",
        "module": "MOD-DOC",
        "rf": "RF-DOC-10",
        "status": "Parcial",
        "location": "app/Http/Controllers/Api/ExpedienteController.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/ExpedienteDocumentalTest.php",
        "commit": COMMITS["backend"],
        "prompt": "R-002",
        "prompt_text": "Búsqueda por código y asunto en tiempo real",
        "notes": "Falta filtro por unidad (RF-DOC-10)",
    },
    {
        "hu": "HU-DOC-07",
        "title": "Catálogo tipos documentales",
        "module": "MOD-DOC",
        "rf": "RF-DOC-11",
        "status": "Implementado",
        "location": "app/Services/Documentaria/TipoDocumentalService.php",
        "version": "v1.0",
        "modified": "2026-06-26",
        "test": "tests/Feature/ExpedienteDocumentalTest.php",
        "commit": COMMITS["integraciones"],
        "prompt": "R-002",
        "prompt_text": "CRUD tipos filtrados por unidad emisora",
        "notes": "TiposDocumentalesPage.vue; PA-29",
    },
    {
        "hu": "HU-DOC-08",
        "title": "Firma digital y sello",
        "module": "MOD-DOC",
        "rf": "RF-DOC-12, RF-DOC-13",
        "status": "Implementado",
        "location": "app/Services/Documentaria/FirmaService.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/ExpedienteDocumentalTest.php",
        "commit": COMMITS["backend"],
        "prompt": "D-002",
        "prompt_text": "Firma HMAC-SHA256 + sello PDF institucional",
        "notes": "ADR-004; no PKI externa",
    },
    {
        "hu": "HU-PAT-01",
        "title": "Inventario patrimonial (Patrimonio / UTIS)",
        "module": "MOD-PAT-TI",
        "rf": "RF-PAT-01, RF-PAT-02, RF-PAT-03, RF-PAT-04",
        "status": "Implementado",
        "location": "app/Services/Patrimonio/EquipoService.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/PatrimonioTiTest.php",
        "commit": COMMITS["backend"],
        "prompt": "D-001",
        "prompt_text": "Inventario TI; vista completa Patrimonio vs parcial UTIS",
        "notes": "InventarioPage.vue",
    },
    {
        "hu": "HU-PAT-02",
        "title": "Incidencias técnicas",
        "module": "MOD-PAT-TI",
        "rf": "RF-PAT-06, RF-PAT-07",
        "status": "Implementado",
        "location": "app/Services/Patrimonio/IncidenciaService.php",
        "version": "v1.0",
        "modified": "2026-06-26",
        "test": "tests/Feature/PatrimonioTiTest.php",
        "commit": COMMITS["integraciones"],
        "prompt": "R-002",
        "prompt_text": "Ciclo abierta → en_atencion → cerrada",
        "notes": "IncidenciasPage.vue",
    },
    {
        "hu": "HU-PAT-03",
        "title": "ML Random Forest predictivo",
        "module": "MOD-PAT-TI",
        "rf": "RF-PAT-08",
        "status": "Implementado",
        "location": "app/Services/Patrimonio/MlPredictionService.php",
        "version": "v1.0",
        "modified": "2026-06-26",
        "test": "tests/Feature/PatrimonioTiTest.php",
        "commit": COMMITS["integraciones"],
        "prompt": "D-002",
        "prompt_text": "Predicción falla vía FastAPI o fallback local",
        "notes": "ADR-002; sgmi:ml-predict diario",
    },
    {
        "hu": "HU-PAT-04",
        "title": "Alertas preventivas (semáforo)",
        "module": "MOD-PAT-TI",
        "rf": "RF-PAT-09, RF-PAT-10",
        "status": "Implementado",
        "location": "app/Http/Controllers/Api/MlPrediccionController.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "tests/Feature/PatrimonioTiTest.php",
        "commit": COMMITS["backend"],
        "prompt": "D-003-v2",
        "prompt_text": "Semáforo verde/amarillo/rojo y equipos críticos",
        "notes": "SemaforoPage.vue; SemaphoreChart.vue",
    },
    {
        "hu": "HU-PAT-05",
        "title": "Fichas técnicas y de mantenimiento",
        "module": "MOD-PAT-TI",
        "rf": "RF-PAT-05",
        "status": "Implementado",
        "location": "app/Services/Patrimonio/FichaService.php",
        "version": "v1.0",
        "modified": "2026-06-12",
        "test": "—",
        "commit": COMMITS["backend"],
        "prompt": "D-001",
        "prompt_text": "Ficha técnica y mantenimiento por equipo",
        "notes": "EquipoDetallePage.vue; sin test Feature dedicado",
    },
    {
        "hu": "HU-DASH-01",
        "title": "Eficiencia en tramitación (sin SLA)",
        "module": "MOD-DASH",
        "rf": "RF-DASH-01, RF-DASH-02, RF-DASH-05",
        "status": "Implementado",
        "location": "app/Services/Dashboard/DashboardService.php",
        "version": "v1.0",
        "modified": "2026-06-26",
        "test": "tests/Feature/DashboardTest.php",
        "commit": COMMITS["integraciones"],
        "prompt": "D-003-v2",
        "prompt_text": "KPIs tramitación y cuellos de botella por gerencia",
        "notes": "DashboardPage.vue; alcance por rol",
    },
    {
        "hu": "HU-DASH-02",
        "title": "Consolidación alertas TI",
        "module": "MOD-DASH",
        "rf": "RF-DASH-03",
        "status": "Implementado",
        "location": "app/Services/Dashboard/DashboardService.php",
        "version": "v1.0",
        "modified": "2026-06-26",
        "test": "tests/Feature/DashboardTest.php",
        "commit": COMMITS["integraciones"],
        "prompt": "D-003-v2",
        "prompt_text": "Dashboard estratégico con semáforo e incidencias",
        "notes": "DashboardEstrategicoPage.vue",
    },
    {
        "hu": "HU-DASH-03",
        "title": "Ejecución presupuestal (SIAF restringido)",
        "module": "MOD-DASH",
        "rf": "RF-DASH-04, RF-DASH-06",
        "status": "Implementado",
        "location": "app/Services/Dashboard/DashboardService.php",
        "version": "v1.0",
        "modified": "2026-06-26",
        "test": "tests/Feature/DashboardTest.php",
        "commit": COMMITS["integraciones"],
        "prompt": "D-002",
        "prompt_text": "Bloque SIAF solo lectura con permiso dash.siaf.ver",
        "notes": "Snapshot local; roles FINANZAS_SIAF",
    },
    {
        "hu": "HU-INT-01",
        "title": "Importación patrimonio SIGA vía API",
        "module": "INT",
        "rf": "RI-SIGA-01, RI-SIGA-02, RI-SIGA-04",
        "status": "Implementado",
        "location": "app/Services/Integrations/SigaSyncService.php",
        "version": "v1.0",
        "modified": "2026-06-26",
        "test": "tests/Feature/IntegracionTest.php",
        "commit": COMMITS["integraciones"],
        "prompt": "D-002",
        "prompt_text": "Sync diario y manual patrimonio SIGA; upsert codigo_siga",
        "notes": "SyncSigaPatrimonioJob; IntegracionesPage.vue",
    },
    {
        "hu": "HU-INT-02",
        "title": "Mapeo personal y oficinas SIGA",
        "module": "INT",
        "rf": "RI-SIGA-03",
        "status": "Implementado",
        "location": "app/Jobs/SyncSigaOrganigramaJob.php",
        "version": "v1.0",
        "modified": "2026-06-26",
        "test": "tests/Feature/IntegracionTest.php",
        "commit": COMMITS["integraciones"],
        "prompt": "D-002",
        "prompt_text": "Sync organigrama y personal SIGA a tablas locales",
        "notes": "PersonalSigaReferencia; fixtures JSON",
    },
    {
        "hu": "HU-INT-03",
        "title": "Lectura ejecución SIAF",
        "module": "INT",
        "rf": "RI-SIAF-01, RI-SIAF-02",
        "status": "Implementado",
        "location": "app/Services/Integrations/SiafSyncService.php",
        "version": "v1.0",
        "modified": "2026-06-26",
        "test": "tests/Feature/IntegracionTest.php",
        "commit": COMMITS["integraciones"],
        "prompt": "D-002",
        "prompt_text": "Sync solo lectura ejecución presupuestal SIAF",
        "notes": "SyncSiafEjecucionJob; SiafEjecucionSnapshot",
    },
    {
        "hu": "HU-INT-04",
        "title": "Simuladores SIGA y SIAF",
        "module": "INT",
        "rf": "RI-SIGA-05, RI-SIAF-03",
        "status": "Implementado",
        "location": "app/Integrations/Siga/SigaSimulatorClient.php",
        "version": "v1.0",
        "modified": "2026-06-26",
        "test": "tests/Feature/IntegracionTest.php",
        "commit": COMMITS["integraciones"],
        "prompt": "D-002",
        "prompt_text": "Drivers simulator con fixtures JSON para desarrollo",
        "notes": "config/integrations.php; INTEGRATION_*_DRIVER",
    },
]

HEADER_FILL = PatternFill("solid", fgColor="2A9D8F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")


def func_id(hu: str) -> str:
    return f"F-{hu.replace('HU-', '')}"


def load_trace_commits() -> dict:
    if not COMMITS_JSON.exists():
        return {}
    return json.loads(COMMITS_JSON.read_text(encoding="utf-8"))


def trace_commit(h: dict, registry: dict) -> str:
    entry = registry.get(h["hu"])
    if entry and entry.get("short"):
        return entry["short"]
    return h["commit"]


def style_header_row(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def clear_data_rows(ws, start_row: int = 2) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def fill_functions_sheet(ws) -> None:
    clear_data_rows(ws)
    for idx, h in enumerate(HISTORIAS, start=2):
        ws.cell(row=idx, column=1, value=func_id(h["hu"]))
        ws.cell(row=idx, column=2, value=h["module"])
        ws.cell(row=idx, column=3, value=f"{h['hu']} — {h['title']}")
        ws.cell(row=idx, column=4, value=h["location"])
        ws.cell(row=idx, column=5, value="Equipo SGMI")
        ws.cell(row=idx, column=6, value=h["status"])
        ws.cell(row=idx, column=7, value=h["version"])
        ws.cell(row=idx, column=8, value=datetime.strptime(h["modified"], "%Y-%m-%d"))
        ws.cell(row=idx, column=9, value=h["notes"])
        for col in range(1, 10):
            ws.cell(row=idx, column=col).alignment = WRAP
    style_header_row(ws, 9)
    ws.freeze_panes = "A2"
    for letter, width in {"A": 14, "B": 14, "C": 42, "D": 48, "E": 14, "F": 14, "G": 10, "H": 16, "I": 40}.items():
        ws.column_dimensions[letter].width = width


def fill_prompts_sheet(ws, registry: dict) -> None:
    clear_data_rows(ws)
    prompt_rows = []
    for h in HISTORIAS:
        p = PROMPTS[h["prompt"]]
        sha = trace_commit(h, registry)
        prompt_rows.append(
            {
                "id": f"{p['id']}-{h['hu']}",
                "func": func_id(h["hu"]),
                "text": h["prompt_text"],
                "version": p["version"],
                "technique": p["technique"],
                "status": "Aprobado" if h["status"] == "Implementado" else "En revisión",
                "expected": h["title"],
                "evidence": f"commit {sha} | impl {h['commit']} | {p['file']}",
            }
        )

    for idx, row in enumerate(prompt_rows, start=2):
        ws.cell(row=idx, column=1, value=row["id"])
        ws.cell(row=idx, column=2, value=row["func"])
        ws.cell(row=idx, column=3, value=row["text"])
        ws.cell(row=idx, column=4, value=row["version"])
        ws.cell(row=idx, column=5, value=row["technique"])
        ws.cell(row=idx, column=6, value=row["status"])
        ws.cell(row=idx, column=7, value=row["expected"])
        ws.cell(row=idx, column=8, value=row["evidence"])
        for col in range(1, 9):
            ws.cell(row=idx, column=col).alignment = WRAP

    style_header_row(ws, 8)
    ws.freeze_panes = "A2"
    for letter, width in {"A": 22, "B": 14, "C": 55, "D": 12, "E": 16, "F": 14, "G": 36, "H": 50}.items():
        ws.column_dimensions[letter].width = width


def fill_link_sheet(ws, registry: dict) -> None:
    clear_data_rows(ws)
    for idx, h in enumerate(HISTORIAS, start=2):
        p = PROMPTS[h["prompt"]]
        sha = trace_commit(h, registry)
        ws.cell(row=idx, column=1, value=f"{h['hu']} — {h['title']}")
        ws.cell(row=idx, column=2, value=h["rf"])
        ws.cell(row=idx, column=3, value=h["prompt_text"])
        ws.cell(row=idx, column=4, value=p["version"])
        ws.cell(row=idx, column=5, value=h["location"])
        ws.cell(row=idx, column=6, value=h["version"])
        ws.cell(row=idx, column=7, value=h["test"])
        ws.cell(row=idx, column=8, value=h["status"])
        ws.cell(row=idx, column=9, value=f"commit {sha}")
        ws.cell(row=idx, column=10, value=f"{h['notes']} | impl: {h['commit']}")
        for col in range(1, 11):
            ws.cell(row=idx, column=col).alignment = WRAP

    style_header_row(ws, 10)
    ws.freeze_panes = "A2"
    for letter, width in {
        "A": 42,
        "B": 28,
        "C": 48,
        "D": 12,
        "E": 48,
        "F": 10,
        "G": 38,
        "H": 14,
        "I": 18,
        "J": 40,
    }.items():
        ws.column_dimensions[letter].width = width


def add_summary_sheet(wb, registry: dict) -> None:
    if "Resumen SGMI" in wb.sheetnames:
        del wb["Resumen SGMI"]
    ws = wb.create_sheet("Resumen SGMI", 0)
    ws["A1"] = "Matriz de doble entrada — SGMI-MPA"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"] = "Proyecto: Sistema de Gestión Municipal Integral"
    ws["A4"] = "Metodología: Prompt-Centered SDLC v1.2"
    ws["A6"] = "Total historias de usuario"
    ws["B6"] = len(HISTORIAS)
    ws["A7"] = "Implementadas"
    ws["B7"] = sum(1 for h in HISTORIAS if h["status"] == "Implementado")
    ws["A8"] = "Parciales"
    ws["B8"] = sum(1 for h in HISTORIAS if h["status"] == "Parcial")
    ws["A9"] = "Pendientes"
    ws["B9"] = sum(1 for h in HISTORIAS if h["status"] == "Pendiente")
    ws["A11"] = "Commits trazabilidad (1 por HU)"
    row = 12
    for h in HISTORIAS:
        sha = trace_commit(h, registry)
        ws.cell(row=row, column=1, value=h["hu"])
        ws.cell(row=row, column=2, value=sha)
        row += 1
    ws["A" + str(row + 1)] = "Commits implementación históricos"
    row += 2
    for label, sha in [
        ("Estructura inicial + requisitos", COMMITS["init"]),
        ("Infraestructura frontend Vue 3", COMMITS["frontend"]),
        ("Backend MVP", COMMITS["backend"]),
        ("Integraciones SIGA/SIAF + ML", COMMITS["integraciones"]),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=sha)
        row += 1
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12


def main() -> None:
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"No se encontró la plantilla: {TEMPLATE}")

    registry = load_trace_commits()

    wb = load_workbook(TEMPLATE)
    fill_functions_sheet(wb["Trazabilidad de funciones"])
    fill_prompts_sheet(wb["Trazabilidad del prompt"], registry)
    fill_link_sheet(wb["Enlace función-prompt"], registry)
    add_summary_sheet(wb, registry)

    OUTPUT_PROJECT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PROJECT)
    shutil.copy2(OUTPUT_PROJECT, OUTPUT_DOWNLOADS)

    print(f"Matriz generada: {OUTPUT_PROJECT}")
    print(f"Copia en Downloads: {OUTPUT_DOWNLOADS}")
    print(f"Historias documentadas: {len(HISTORIAS)}")
    print(f"  Implementado: {sum(1 for h in HISTORIAS if h['status'] == 'Implementado')}")
    print(f"  Parcial: {sum(1 for h in HISTORIAS if h['status'] == 'Parcial')}")


if __name__ == "__main__":
    main()
