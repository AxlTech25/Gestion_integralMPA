#!/usr/bin/env python3
"""Analyze the double-entry matrix Excel template."""
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook

path = Path(r"c:\Users\INTEL\Downloads\PLANTILLA - MATRIZ DE DOBLE ENTRADA.xlsx")
wb = load_workbook(path, data_only=True)

print("=== HOJAS ===")
for name in wb.sheetnames:
    print(f"  - {name}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== HOJA: {sheet_name} ===")
    print(f"Dimensiones: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

    print("Contenido (primeras 35 filas):")
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(35, ws.max_row), values_only=False), 1
    ):
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                s = repr(v)
                if len(s) > 80:
                    s = s[:77] + "..."
                vals.append(f"{cell.coordinate}={s}")
        if vals:
            print(f"  Fila {row_idx}: " + " | ".join(vals))

    if ws.merged_cells.ranges:
        merged = [str(m) for m in ws.merged_cells.ranges]
        print(f"Celdas combinadas ({len(merged)}):")
        for m in merged[:25]:
            print(f"  {m}")
        if len(merged) > 25:
            print(f"  ... y {len(merged) - 25} mas")

ws = wb[wb.sheetnames[0]]
print("\n=== ENCABEZADOS / ESTRUCTURA COMPLETA FILA 1-10 ===")
for row_idx in range(1, min(11, ws.max_row + 1)):
    row_vals = []
    for col_idx in range(1, min(ws.max_column + 1, 20)):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            row_vals.append(f"Col{col_idx}({cell.coordinate})={repr(cell.value)[:100]}")
    if row_vals:
        print(f"Fila {row_idx}: " + " | ".join(row_vals))

print("\n=== COLUMNAS CON DATOS (todas las filas, hoja 1) ===")
for col_idx in range(1, ws.max_column + 1):
    col_letter = ws.cell(row=1, column=col_idx).column_letter
    non_empty = []
    for row_idx in range(1, ws.max_row + 1):
        v = ws.cell(row=row_idx, column=col_idx).value
        if v is not None:
            non_empty.append((row_idx, str(v)[:60]))
    if non_empty:
        print(f"\nColumna {col_letter} ({col_idx}) - {len(non_empty)} celdas con datos:")
        for r, v in non_empty[:15]:
            print(f"  R{r}: {v}")
        if len(non_empty) > 15:
            print(f"  ... y {len(non_empty) - 15} mas")
